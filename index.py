import os
import openai
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
import requests
from google.cloud import vision
from flask import Flask, request, render_template, send_file, redirect, url_for, flash
from flask_socketio import SocketIO, emit
from werkzeug.utils import secure_filename
import time
import concurrent.futures
import logging
import re
from azure.cognitiveservices.vision.computervision import ComputerVisionClient
from msrest.authentication import CognitiveServicesCredentials

app = Flask(__name__)
app.secret_key = 'supersecretkey'
app.config['UPLOAD_FOLDER'] = r'C:\xampp\htdocs\py\images'
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # Set maximum upload size to 100MB

socketio = SocketIO(app)

# Load environment variables
load_dotenv()

# Set OpenAI API key
openai.api_key = os.getenv('OPENAI_API_KEY')

# Set up Google Cloud Vision client
os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = r'C:\xampp\htdocs\py\personaldrivers-497ceba0ad12.json'
vision_client = vision.ImageAnnotatorClient()

# Set up Azure Computer Vision client
azure_endpoint = "https://dilshad1.cognitiveservices.azure.com/"
azure_api_key = "e4a556b281eb4a87a927fdf7e4d97261"
azure_client = ComputerVisionClient(azure_endpoint, CognitiveServicesCredentials(azure_api_key))

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


def validate_image(file):
    allowed_extensions = {'png', 'jpg', 'jpeg', 'gif'}
    return '.' in file.filename and file.filename.rsplit('.', 1)[1].lower() in allowed_extensions


def analyze_image(image_path):
    with open(image_path, 'rb') as image_file:
        content = image_file.read()
        image = vision.Image(content=content)

    descriptions = []
    detailed_description = []
    google_objects = set()
    azure_objects = set()

    # Google Vision Object Detection
    try:
        logging.info("Performing Google Vision Object Detection")
        object_response = vision_client.object_localization(image=image)
        objects = object_response.localized_object_annotations
        google_objects = {obj.name.lower() for obj in objects}
        logging.info(f"Google Detected Objects: {google_objects}")
        object_description = 'Google Objects: ' + ', '.join([f"{obj.name} (confidence: {obj.score:.2f})" for obj in objects])
        descriptions.append(object_description)
        detailed_description.append('Google Objects: ' + ', '.join([f"{obj.name} (confidence: {obj.score:.2f})" for obj in objects]))
    except Exception as e:
        logging.error(f"Google Object Localization Error: {e}")
        descriptions.append(f"Google Object Localization Error: {e}")

    # Azure Vision Object Detection
    try:
        logging.info("Performing Azure Vision Object Detection")
        with open(image_path, 'rb') as azure_image_file:
            azure_image_analysis = azure_client.analyze_image_in_stream(azure_image_file, visual_features=["objects"])
        azure_objects = {obj.object_property.lower() for obj in azure_image_analysis.objects}
        logging.info(f"Azure Detected Objects: {azure_objects}")
        object_description = 'Azure Objects: ' + ', '.join([f"{obj.object_property} (confidence: {obj.confidence:.2f})" for obj in azure_image_analysis.objects])
        descriptions.append(object_description)
        detailed_description.append('Azure Objects: ' + ', '.join([f"{obj.object_property} (confidence: {obj.confidence:.2f})" for obj in azure_image_analysis.objects]))
    except Exception as e:
        logging.error(f"Azure Object Detection Error: {e}")
        descriptions.append(f"Azure Object Detection Error: {e}")

    # Find intersection of objects detected by both Google and Azure
    common_objects = google_objects.intersection(azure_objects)
    if common_objects:
        common_description = 'Common Objects: ' + ', '.join(common_objects)
        descriptions.append(common_description)
        detailed_description.append(common_description)
    else:
        logging.info("No common objects detected by both Google and Azure.")

    # Combine all descriptions into one
    combined_description = ' | '.join(descriptions)
    detailed_combined_description = ' | '.join(detailed_description)
    return combined_description, detailed_combined_description


def generate_metadata_for_image(image_description, detailed_image_description, image_file_name, user_prompt):
    url = "https://api.openai.com/v1/chat/completions"

    # Enhanced prompt with more specific instructions and length constraints
    prompt_template = (
        f"You are an SEO expert specializing in generating image metadata for online listings. "
        f"Your task is to write three fields: **Alt Text**, **Caption**, and **Description**. These will be used for image optimization in search engines. "
        f"Follow these specific instructions carefully:\n\n"
        
        f"1. **Alt Text (max 125 characters)**: Describe the image as clearly and naturally as possible, without including any marketing or promotional language. Focus only on what can be visually observed. "
        f"   Example: 'A white car parked by a scenic mountain road on a bright sunny day.'\n\n"
        
        f"2. **Caption (max 120 characters)**: Create an engaging and brief summary of the image that would appeal to users on social media. It should be more engaging than the Alt Text but still succinct. "
        f"   Example: 'Scenic road trip adventure with a white car parked near stunning mountain views.'\n\n"
        
        f"3. **Description (max 160 characters)**: Write a more detailed description that includes relevant SEO keywords and potentially a soft Call to Action (CTA) if relevant. Focus on providing context or use cases."
        f"   Example: 'Planning a scenic road trip? Hire our professional drivers for safe and convenient car transport across beautiful landscapes.'\n\n"
        
        f"User's provided context: {user_prompt}.\n"
        f"Here is the general description of the image: {image_description}.\n"
        f"Here is the detailed description of the image: {detailed_image_description}."
    )

    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {os.getenv('OPENAI_API_KEY')}"
    }

    def get_metadata(prompt):
        payload = {
            "model": "gpt-4o",
            "messages": [
                {"role": "system", "content": "You are an SEO expert."},
                {"role": "user", "content": prompt}
            ],
            "temperature": 0.3,  # Lower temperature for more precise and factual output
            "max_tokens": 500,
            "top_p": 1,
            "frequency_penalty": 2,
            "presence_penalty": 2
        }

        response = requests.post(url, json=payload, headers=headers)
        if response.status_code != 200:
            logging.error(f"Error: {response.status_code}, {response.text}")
            return None
        return response.json()['choices'][0]['message']['content'].strip()

    def clean_text(text):
        return re.sub(r'[^a-zA-Z0-9\s,.!]', '', text).strip()

    def parse_generated_texts(generated_texts):
        alt_text = ""
        caption = ""
        description = ""
        try:
            # Stricter parsing with explicit keyword search
            if "Alt Text:" in generated_texts:
                alt_text = clean_text(generated_texts.split("Alt Text:")[1].split("Caption:")[0].strip())
            if "Caption:" in generated_texts:
                caption = clean_text(generated_texts.split("Caption:")[1].split("Description:")[0].strip())
            if "Description:" in generated_texts:
                description = clean_text(generated_texts.split("Description:")[1].split('\n')[0].strip())
        except Exception as e:
            logging.error(f"Error parsing metadata: {e}")
        return alt_text, caption, description

    # Fetch the metadata from OpenAI
    generated_texts = get_metadata(prompt_template)
    if not generated_texts:
        return image_file_name, "Default Alt Text", "Default Caption", "Default Description"

    # Parse and clean the generated content
    alt_text, caption, description = parse_generated_texts(generated_texts)
    
    # Retry if any part of the metadata is missing
    if not alt_text or not caption or not description:
        logging.info("Retrying to fetch missing metadata fields.")
        while not alt_text or not caption or not description:
            missing_fields_prompt = prompt_template + f" Missing fields: {'Alt Text' if not alt_text else ''}{', Caption' if not caption else ''}{', Description' if not description else ''}."
            generated_texts = get_metadata(missing_fields_prompt)
            if generated_texts:
                new_alt_text, new_caption, new_description = parse_generated_texts(generated_texts)
                alt_text = alt_text or new_alt_text
                caption = caption or new_caption
                description = description or new_description

    return image_file_name, alt_text, caption, description


# Function to store data in an Excel file with images
def store_data_in_excel(data, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Image Metadata"
    ws.append(['Image', 'Image Name', 'Alt Text', 'Caption', 'Description'])

    for row in data:
        image_path, image_name, alt_text, caption, description = row
        img = ExcelImage(image_path)
        img.width = 100  # Set image width
        img.height = 100  # Set image height

        # Set row height and column width
        row_num = ws.max_row + 1
        ws.row_dimensions[row_num].height = 80
        col_letter = get_column_letter(1)
        ws.column_dimensions[col_letter].width = 20

        ws.append([None, image_name, alt_text, caption, description])
        ws.add_image(img, f'A{ws.max_row}')

    wb.save(output_file)


@app.route('/')
def upload_form():
    return render_template('upload.html')


@app.route('/upload', methods=['POST'])
def upload_image():
    if 'files[]' not in request.files or 'prompt' not in request.form:
        flash('No file or prompt part')
        return redirect(request.url)

    user_prompt = request.form['prompt']
    files = request.files.getlist('files[]')
    if not files or files[0].filename == '':
        flash('No selected files')
        return redirect(request.url)

    if len(files) > 20:
        flash('You can upload a maximum of 20 images at a time.')
        return redirect(request.url)

    data = []
    total_files = len(files)

    with concurrent.futures.ThreadPoolExecutor() as executor:
        future_to_file = {executor.submit(process_image, file, user_prompt, i, total_files): file for i, file in enumerate(files)}
        for future in concurrent.futures.as_completed(future_to_file):
            file = future_to_file[future]
            try:
                result = future.result()
                if result:
                    data.append(result)
            except Exception as e:
                logging.error(f"Exception occurred: {e}")

    output_file = 'image_metadata.xlsx'
    store_data_in_excel(data, output_file)
    return send_file(output_file, as_attachment=True)


def process_image(file, user_prompt, index, total_files):
    if file and validate_image(file):
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        image_description, detailed_image_description = analyze_image(file_path)
        image_file_name, alt_text, caption, description = generate_metadata_for_image(image_description, detailed_image_description, filename, user_prompt)
        
        # Emit progress to the client
        progress = (index + 1) / total_files * 100
        socketio.emit('progress', {'progress': progress}, namespace='/')
        time.sleep(0.1)  # Simulate processing time for demo purposes
        
        return [file_path, image_file_name, alt_text, caption, description]
    else:
        logging.warning(f"Invalid file type: {file.filename}")
        return None


if __name__ == "__main__":
    socketio.run(app, debug=True, port=5002)
